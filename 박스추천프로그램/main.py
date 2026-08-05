import os
import codecs
import datetime
import itertools
import urllib.parse
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import openpyxl
import pandas as pd
import fitz  # PyMuPDF
from sqlalchemy import create_engine, text
from sqlalchemy.exc import SQLAlchemyError
from tkinterdnd2 import TkinterDnD, DND_FILES
from dotenv import load_dotenv

load_dotenv()

# --- 3D 렌더링을 위한 Matplotlib 임포트 ---
try:
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

# -----------------------------------------------------------------------------
# 1. 시스템 설정 및 에러 핸들러
# -----------------------------------------------------------------------------

DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT", "5432")
DB_USER = os.getenv("DB_USER", "postgres")
DB_PASSWORD = os.getenv("DB_PASSWORD")
DB_NAME = os.getenv("DB_NAME", "encluscm")

KOREAN_FONT_PATH = "C:/Windows/Fonts/malgun.ttf"

BOX_CANDIDATES = [
    ("봉투", 9, 6, 1), ("B-16", 16, 12, 9), ("B-21", 17, 13, 12),
    ("B-54", 24, 12, 10.3), ("B-34", 20, 15, 10), ("B-53", 23, 20, 9),
    ("A-23", 31, 21, 70), ("C-8", 22, 16, 14), ("B-70", 27, 21, 10),
    ("B-56", 24, 20, 12), ("C-14", 23, 20, 18), ("B-89", 30.5, 21.5, 16),
    ("A-13-1", 28, 20, 20), ("B-134", 40, 25, 15), ("C-133", 36, 36, 12),
    ("C-148", 40, 25, 20), ("C-153", 40, 31, 21), ("B-138", 40, 31, 31)
]

# -----------------------------------------------------------------------------
# 2. 데이터베이스 엔진 (싱글톤 패턴 적용)
# -----------------------------------------------------------------------------
_DB_ENGINE = None

def get_db_engine():
    global _DB_ENGINE
    if _DB_ENGINE is not None:
        return _DB_ENGINE
    
    try:
        safe_password = urllib.parse.quote_plus(DB_PASSWORD)
        connection_url = f"postgresql+psycopg2://{DB_USER}:{safe_password}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
        _DB_ENGINE = create_engine(connection_url, client_encoding='utf8', pool_pre_ping=True)
        return _DB_ENGINE
    except Exception as e:
        print(f"DB Engine Error: {e}")
        return None

# -----------------------------------------------------------------------------
# 3. 데이터베이스 CRUD 헬퍼 함수
# -----------------------------------------------------------------------------
def load_combination_boxes():
    combos = {}
    engine = get_db_engine()
    if not engine: return combos
    try:
        with engine.connect() as conn:
            df = pd.read_sql("SELECT combo_key, box_name FROM box_combinations;", conn)
            for _, row in df.iterrows():
                combos[str(row['combo_key'])] = str(row['box_name'])
    except SQLAlchemyError as e:
        print(f"Error loading combos: {e}")
    return combos

def save_or_update_combo(combo_key, box_name):
    engine = get_db_engine()
    if not engine: return
    try:
        with engine.begin() as conn:
            sql = text("""
            INSERT INTO box_combinations (combo_key, box_name, updated_at) 
            VALUES (:key, :name, CURRENT_TIMESTAMP)
            ON CONFLICT (combo_key) 
            DO UPDATE SET box_name = EXCLUDED.box_name, updated_at = CURRENT_TIMESTAMP;
            """)
            conn.execute(sql, {"key": combo_key, "name": box_name})
    except SQLAlchemyError as e:
        print(f"Error saving combo: {e}")

def delete_combo_from_db(combo_key):
    engine = get_db_engine()
    if not engine: return
    try:
        with engine.begin() as conn:
            sql = text("DELETE FROM box_combinations WHERE combo_key = :key;")
            conn.execute(sql, {"key": combo_key})
    except SQLAlchemyError as e:
        print(f"Error deleting combo: {e}")

def load_products():
    products = {}
    engine = get_db_engine()
    if not engine: return products
    try:
        with engine.connect() as conn:
            df = pd.read_sql("SELECT product_code, width, depth, height FROM products;", conn)
            for _, row in df.iterrows():
                products[str(row['product_code'])] = (float(row['width']), float(row['depth']), float(row['height']))
    except SQLAlchemyError as e:
        print(f"Error loading products: {e}")
    return products

def save_or_update_product(code, dims):
    engine = get_db_engine()
    if not engine: return
    try:
        with engine.begin() as conn:
            sql = text("""
            INSERT INTO products (product_code, width, depth, height, updated_at) 
            VALUES (:code, :w, :d, :h, CURRENT_TIMESTAMP)
            ON CONFLICT (product_code) 
            DO UPDATE SET width = EXCLUDED.width, depth = EXCLUDED.depth, height = EXCLUDED.height, updated_at = CURRENT_TIMESTAMP;
            """)
            conn.execute(sql, {"code": code, "w": dims[0], "d": dims[1], "h": dims[2]})
    except SQLAlchemyError as e:
        print(f"Error saving product: {e}")

def delete_product_from_db(code):
    engine = get_db_engine()
    if not engine: return
    try:
        with engine.begin() as conn:
            conn.execute(text("DELETE FROM products WHERE product_code = :code;"), {"code": code})
    except SQLAlchemyError as e:
        print(f"Error deleting product: {e}")

# -----------------------------------------------------------------------------
# 4. 박스 크기 연산 알고리즘 (Best-Fit 3D Bin Packing - 위로 쌓기 최적화)
# -----------------------------------------------------------------------------
def get_rotations(dims):
    """상품을 가로, 세로, 높이로 돌려보는 6가지 방향 반환"""
    return list(set(itertools.permutations(dims)))

def try_packing_in_box(items, box_w, box_d, box_h):
    """
    제한된 박스 크기(W, D, H) 안에 상품이 겹치지 않고 모두 들어가는지 시뮬레이션.
    성공 시 True와 함께 배치 좌표 반환.
    """
    occupied_spaces = []
    
    for item in items:
        if len(item) >= 4:
            code, iw_orig, id_orig, ih_orig = item[0], item[1], item[2], item[3]
        else:
            code, iw_orig, id_orig, ih_orig = "상품", item[0], item[1], item[2]
            
        rotations = get_rotations((iw_orig, id_orig, ih_orig))
        
        potentials = [(0, 0, 0)]
        for space in occupied_spaces:
            ox, oy, oz, ow, od, oh = space[:6]
            # 기존 물건의 주변 빈 공간을 후보 좌표로 추가
            potentials.extend([
                (ox + ow, oy, oz), (ox, oy + od, oz), (ox, oy, oz + oh),
                (0, oy + od, oz), (ox + ow, 0, oz), (ox, oy, 0),
                (0, 0, oz + oh)
            ])
            
        best_placement = None
        best_score = float('inf')
        
        for px, py, pz in set(potentials):
            for iw, id_, ih in rotations:
                # 박스 범위 초과 검사 (상자 밖으로 튀어나가면 무시)
                if px + iw > box_w or py + id_ > box_d or pz + ih > box_h:
                    continue
                    
                # 기존에 배치된 물건들과의 충돌(겹침) 검사
                overlap = False
                for space in occupied_spaces:
                    ox, oy, oz, ow, od, oh = space[:6]
                    if not (px + iw <= ox or px >= ox + ow or py + id_ <= oy or py >= oy + od or pz + ih <= oz or pz >= oz + oh):
                        overlap = True
                        break
                        
                if not overlap:
                    # [위로 쌓기 핵심 로직] 
                    # 원점(0,0,0)에 가깝게 붙이되, Z축(높이) 페널티를 줄여 위로 쌓는 것을 선호하도록 만듦
                    # footprint(바닥면적 iw*id)가 크면 페널티를 주어 상품을 세우도록 유도함
                    score = (px**2 + py**2 + (pz * 0.8)**2) + (iw * id_) * 0.01
                    
                    if score < best_score:
                        best_score = score
                        best_placement = (px, py, pz, iw, id_, ih, code)
                        
        if best_placement:
            occupied_spaces.append(best_placement)
        else:
            return False, [] # 하나라도 상자에 못 들어가면 이 상자는 실패
            
    return True, occupied_spaces

def find_best_box_and_layout(items, box_list):
    """모든 박스를 탐색하여 가장 작고 효율적인 상자를 추천"""
    if not items:
        return "상품이 없습니다.", None, []
        
    # 최적화: 물리적으로 부피가 모자라거나, 변의 길이가 초과하는 상자 쳐내기
    total_volume = sum(x[1]*x[2]*x[3] if len(x)>=4 else x[0]*x[1]*x[2] for x in items)
    max_item_dim = max(max(x[1:4]) if len(x)>=4 else max(x[:3]) for x in items)
    
    # 상자를 부피가 작은 순서대로 정렬 (가장 작은 상자부터 씌워봄)
    sorted_boxes = sorted(box_list, key=lambda b: b[1] * b[2] * b[3])
    
    # 여러 정렬 기준(다단계 최적화)으로 시도
    heuristics = [
        lambda x: x[1]*x[2]*x[3] if len(x)>=4 else x[0]*x[1]*x[2],           # 부피가 큰 것부터
        lambda x: max(x[1:4]) if len(x)>=4 else max(x[:3])                    # 길이가 긴 것부터
    ]
    
    for name, b_w, b_d, b_h in sorted_boxes:
        if (b_w * b_d * b_h) < total_volume: continue
        if max_item_dim > max(b_w, b_d, b_h): continue
            
        # 상자의 방향도 돌려가며 팩킹 시도
        box_rotations = get_rotations((b_w, b_d, b_h))
        for box_rot in box_rotations:
            for heuristic in heuristics:
                sorted_items = sorted(items, key=heuristic, reverse=True)
                success, layout = try_packing_in_box(sorted_items, box_rot[0], box_rot[1], box_rot[2])
                if success:
                    return name, box_rot, layout # 성공하면 즉시 반환 (이미 부피순 정렬되어 있으므로 최적)
                    
    return "⚠️ 규격초과(더 큰 상자 필요)", None, []

# -----------------------------------------------------------------------------
# 5. GUI 메인 애플리케이션
# -----------------------------------------------------------------------------
class IntegratedBoxApp:
    def __init__(self, root):
        self.root = root
        self.root.title("박스크기추천 통합 시스템")
        self.root.geometry("900x850")
        
        engine = get_db_engine()
        if engine:
            try:
                with engine.connect():
                    pass 
            except Exception as e:
                messagebox.showerror("DB 연결 에러", f"원격 DB({DB_HOST})에 연결할 수 없습니다.\n네트워크 및 보안 설정을 확인하세요.\n\n{str(e)}")
        
        self.dropped_excel = None
        self.dropped_pdf = None
        self.dropped_txt = None
        
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True)
        
        self.create_tab1()  
        self.create_tab2()  
        self.create_tab3()  
        self.create_tab4()  
        self.create_tab5()  

    def create_tab1(self):
        tab1 = tk.Frame(self.notebook, bg="#f5f5f5")
        self.notebook.add(tab1, text=" 1번 탭 (주문데이터 처리) ")
        
        self.drop_zone_excel = tk.Label(
            tab1, 
            text="\n-------------------------------------------------------------\n\n 여기에 오늘자 [주문데이터 엑셀 파일 (.xls, .xlsx)]을\n마우스로 끌어다 놓으세요 (Drag & Drop) \n\n-------------------------------------------------------------\n",
            font=("맑은 고딕", 12, "bold"), bg="#efebe9", fg="#4e342e", bd=2, relief="solid"
        )
        self.drop_zone_excel.pack(padx=30, pady=20, fill="x")
        self.drop_zone_excel.drop_target_register(DND_FILES)
        self.drop_zone_excel.dnd_bind('<<Drop>>', self.handle_dropped_excel)
        
        self.lbl_status_excel = tk.Label(tab1, text="주문 엑셀 파일 대기 중...", font=("맑은 고딕", 10, "bold"), bg="#f5f5f5", fg="#ff5722")
        self.lbl_status_excel.pack(pady=5)
        
        self.btn_run_excel = tk.Button(
            tab1, text="📊 주문 데이터 분석 및 박스 추천결과 생성 (DB 기준 실시간 연산)", 
            font=("맑은 고딕", 12, "bold"), bg="#795548", fg="white", state="disabled", command=self.process_excel_matching
        )
        self.btn_run_excel.pack(padx=30, pady=5, fill="x")
        
        result_frame = tk.LabelFrame(tab1, text=" 🖥️ 1번 탭 실시간 콘솔 로그 ")
        result_frame.pack(padx=30, pady=15, fill="both", expand=True)
        self.txt_result_excel = tk.Text(result_frame, font=("맑은 고딕", 10), bg="white", fg="#222222")
        self.txt_result_excel.pack(fill="both", expand=True, padx=10, pady=10)

    def handle_dropped_excel(self, event):
        raw_data = event.data
        files = [f.strip('{}').strip('"') for f in (raw_data.split('} {') if raw_data.startswith('{') else raw_data.split())]
        for file_path in files:
            ext = os.path.splitext(file_path)[-1].lower()
            if ext in ['.xls', '.xlsx']:
                self.dropped_excel = file_path
                break
                
        self.txt_result_excel.delete("1.0", tk.END)
        if self.dropped_excel:
            self.txt_result_excel.insert(tk.END, f"📥 인식된 엑셀 파일:\n   {os.path.basename(self.dropped_excel)}\n")
            self.lbl_status_excel.configure(text="✅ 엑셀 파일 로드 완료! 분석을 시작할 수 있습니다.", fg="#4caf50")
            self.btn_run_excel.configure(state="normal")
        else:
            self.lbl_status_excel.configure(text="⚠️ 올바른 엑셀 파일(.xls, .xlsx)을 던져주세요.", fg="#f44336")
            self.btn_run_excel.configure(state="disabled")

    def process_excel_matching(self):
        if not self.dropped_excel: return
        self.txt_result_excel.insert(tk.END, "\n▶ 원격 DB 기반 데이터 동기화 및 추천 연산을 시작합니다...\n")
        today_str = datetime.datetime.now().strftime("%y%m%d")
        dir_name = os.path.dirname(self.dropped_excel)
        output_txt_file = os.path.join(dir_name, f"{today_str}_박스매칭결과.txt")
        
        try:
            df_order = pd.read_excel(self.dropped_excel, dtype={4: str, 7: str})
            
            e_col_name = df_order.columns[4]
            f_col_name = df_order.columns[5]
            h_col_name = df_order.columns[7]
            
            df_order = df_order.dropna(subset=[h_col_name])
            df_order[e_col_name] = df_order[e_col_name].fillna("")
            df_order[f_col_name] = pd.to_numeric(df_order[f_col_name].fillna(1), errors='coerce').astype(int)
            
            df_sorted = df_order.sort_values(by=[h_col_name, e_col_name], ascending=[True, True])
            grouped = df_sorted.groupby(h_col_name)
            
            today_orders = []
            for invoice_num, group in grouped:
                option_list = []
                for _, row in group.iterrows():
                    opt = str(row[e_col_name]).strip()
                    count = int(row[f_col_name])
                    if opt:
                        for _ in range(count): option_list.append(opt)
                if option_list:
                    options_str = ",".join(option_list)
                    today_orders.append((invoice_num, options_str))
            
            box_master_dict = load_combination_boxes()  
            product_master_dict = load_products()       
            
            final_output_lines = []
            fixed_match_count = 0
            algo_match_count = 0
            fail_count = 0
            
            for invoice, opt_str in today_orders:
                items_in_order = [x.strip() for x in opt_str.split(',') if x.strip()]
                lookup_key = "+".join(sorted(items_in_order))
                
                if lookup_key in box_master_dict:
                    matched_box = box_master_dict[lookup_key]
                    fixed_match_count += 1
                else:
                    item_dims_with_id_list = []
                    has_missing_dim = False
                    for item_code in items_in_order:
                        if item_code in product_master_dict:
                            w, d, h = product_master_dict[item_code]
                            item_dims_with_id_list.append((item_code, w, d, h))
                        else:
                            has_missing_dim = True
                            break
                    
                    if not has_missing_dim and item_dims_with_id_list:
                        # 1번 탭: 새로운 위로 쌓기 알고리즘을 사용해 가장 작은 상자 도출
                        recommended_box, _, _ = find_best_box_and_layout(item_dims_with_id_list, BOX_CANDIDATES)
                        matched_box = recommended_box
                        algo_match_count += 1
                    else:
                        matched_box = "매칭실패(신규조합/크기미등록)"
                        fail_count += 1
                        
                final_output_lines.append(f"{invoice}+{opt_str}+{matched_box}")
                
            with open(output_txt_file, "w", encoding="utf-8-sig") as f:
                f.write("\n".join(final_output_lines))
                
            self.txt_result_excel.insert(tk.END, "─" * 40 + "\n")
            self.txt_result_excel.insert(tk.END, f"🎉 원격 DB 기반 매칭 완료!\n📂 결과 파일: {os.path.basename(output_txt_file)}\n")
            messagebox.showinfo("성공", f"DB 분석 완료!\n\n(고정조합: {fixed_match_count}건\n알고리즘 추천: {algo_match_count}건\n매칭 실패: {fail_count}건)")
            self.dropped_txt = output_txt_file
            
        except Exception as e:
            self.txt_result_excel.insert(tk.END, f"❌ 엑셀 분석 중 에러 발생: {str(e)}\n")

    def create_tab2(self):
        tab2 = tk.Frame(self.notebook, bg="#f5f5f5")
        self.notebook.add(tab2, text=" 2번 탭 (PDF 송장 변환) ")
        
        self.drop_zone = tk.Label(
            tab2, 
            text="\n-------------------------------------------------------------\n\n 여기에 [송장 PDF 파일]과 [1번 탭에서 만든 TXT 파일]을\n마우스로 함께 끌어다 놓으세요 (Drag & Drop) \n\n-------------------------------------------------------------\n",
            font=("맑은 고딕", 12, "bold"), bg="#e0f7fa", fg="#006064", bd=2, relief="solid"
        )
        self.drop_zone.pack(padx=30, pady=20, fill="x")
        self.drop_zone.drop_target_register(DND_FILES)
        self.drop_zone.dnd_bind('<<Drop>>', self.handle_dropped_files)
        
        self.lbl_status = tk.Label(tab2, text="파일 대기 중...", font=("맑은 고딕", 10, "bold"), bg="#f5f5f5", fg="#ff5722")
        self.lbl_status.pack(pady=5)
        
        self.btn_run = tk.Button(
            tab2, text="⚡ 엔클루 박스크기 강제 변환 시작", 
            font=("맑은 고딕", 12, "bold"), bg="#008643", fg="white", state="disabled", command=self.process_pdf_invoice
        )
        self.btn_run.pack(padx=30, pady=5, fill="x")
        
        result_frame = tk.LabelFrame(tab2, text=" 🖥️ 2번 탭 실시간 콘솔 로그 ")
        result_frame.pack(padx=30, pady=15, fill="both", expand=True)
        self.txt_result = tk.Text(result_frame, font=("맑은 고딕", 10), bg="white", fg="#222222")
        self.txt_result.pack(fill="both", expand=True, padx=10, pady=10)

    def handle_dropped_files(self, event):
        raw_data = event.data
        files = [f.strip('{}').strip('"') for f in (raw_data.split('} {') if raw_data.startswith('{') else raw_data.split())]
        for file_path in files:
            ext = os.path.splitext(file_path)[-1].lower()
            if ext == '.pdf': self.dropped_pdf = file_path
            elif ext == '.txt': self.dropped_txt = file_path
                
        self.txt_result.delete("1.0", tk.END)
        if self.dropped_pdf: self.txt_result.insert(tk.END, f"  └ 📄 PDF: {os.path.basename(self.dropped_pdf)}\n")
        if self.dropped_txt: self.txt_result.insert(tk.END, f"  └ 📝 TXT: {os.path.basename(self.dropped_txt)}\n")
        
        if self.dropped_pdf and self.dropped_txt:
            self.lbl_status.configure(text="✅ 변환 준비 완료! 버튼을 누르세요.", fg="#4caf50")
            self.btn_run.configure(state="normal")

    def process_pdf_invoice(self):
        if not self.dropped_pdf or not self.dropped_txt: return
        dir_name = os.path.dirname(self.dropped_pdf)
        base_name = os.path.splitext(os.path.basename(self.dropped_pdf))[0]
        output_pdf_file = os.path.join(dir_name, f"{base_name}_완료.pdf")
        
        invoice_box_dict = {}
        try:
            with open(self.dropped_txt, "r", encoding="utf-8-sig") as f:
                lines = f.readlines()
            for line in lines:
                line = line.strip()
                if not line or "+" not in line: continue
                parts = line.split("+")
                if len(parts) >= 3: 
                    invoice_box_dict[parts[0].strip()] = parts[-1].strip()
        except Exception as e:
            messagebox.showerror("텍스트 파일 읽기 오류", str(e))
            return

        try:
            doc = fitz.open(self.dropped_pdf)
            page_updates = {}
            for page_num in range(len(doc)):
                page_text = doc[page_num].get_text("text")
                for invoice, box_size in invoice_box_dict.items():
                    if invoice in page_text:
                        page_updates[page_num] = box_size
                        break

            success_count = 0
            font_path = KOREAN_FONT_PATH if os.path.exists(KOREAN_FONT_PATH) else None
            
            for page_num, box_size in page_updates.items():
                page = doc[page_num]
                target_instances = page.search_for("엔클루")
                if not target_instances: continue
                
                last_instance = max(target_instances, key=lambda inst: inst.y1)
                page.draw_rect(last_instance, color=(1, 1, 1), fill=(1, 1, 1))
                page.insert_text(
                    fitz.Point(last_instance.x0, last_instance.y1 - 3), 
                    f"엔클루   {box_size}", 
                    fontsize=8, 
                    fontfile=font_path, 
                    fontname="ko_font" if font_path else None, 
                    color=(0, 0, 0)
                )
                success_count += 1

            if success_count > 0:
                doc.save(output_pdf_file)
                self.txt_result.insert(tk.END, f"✅ PDF 변환 완료: {success_count}개 페이지 처리됨.\n")
                messagebox.showinfo("성공", f"PDF 내 {success_count}개 페이지 수정 완료!\n저장 경로: {output_pdf_file}")
            else:
                self.txt_result.insert(tk.END, "⚠️ 매칭되는 송장 번호를 PDF에서 찾지 못했습니다.\n")
                
            doc.close()
            
        except Exception as e:
            messagebox.showerror("PDF 처리 오류", str(e))
            self.txt_result.insert(tk.END, f"❌ 에러: {str(e)}\n")

    def create_tab3(self):
        tab3 = tk.Frame(self.notebook)
        self.notebook.add(tab3, text=" 3번 탭 (상품 크기 관리) ")
        
        search_frame = tk.LabelFrame(tab3, text=" 🔍 상품코드 검색 ")
        search_frame.pack(padx=15, pady=5, fill="x")
        self.p_search_var = tk.StringVar()
        self.p_search_var.trace_add("write", lambda *args: self.update_product_tree())
        tk.Entry(search_frame, textvariable=self.p_search_var, width=35).pack(side="left", padx=10, pady=5)
        
        form_frame = tk.LabelFrame(tab3, text=" 상품별 크기 등록 / 수정 (원격 DB 저장) ")
        form_frame.pack(padx=15, pady=5, fill="x")
        
        tk.Label(form_frame, text="옵션코드:").grid(row=0, column=0, padx=5, pady=5)
        self.ent_p_code = tk.Entry(form_frame, width=20); self.ent_p_code.grid(row=0, column=1, padx=5, pady=5)
        tk.Label(form_frame, text="가로:").grid(row=1, column=0, padx=5, pady=5)
        self.ent_p_w = tk.Entry(form_frame, width=10); self.ent_p_w.grid(row=1, column=1, padx=5, pady=5)
        tk.Label(form_frame, text="세로:").grid(row=1, column=2, padx=5, pady=5)
        self.ent_p_d = tk.Entry(form_frame, width=10); self.ent_p_d.grid(row=1, column=3, padx=5, pady=5)
        tk.Label(form_frame, text="높이:").grid(row=1, column=4, padx=5, pady=5)
        self.ent_p_h = tk.Entry(form_frame, width=10); self.ent_p_h.grid(row=1, column=5, padx=5, pady=5)
        
        btn_frame = tk.Frame(form_frame)
        btn_frame.grid(row=2, column=0, columnspan=6, pady=5)
        tk.Button(btn_frame, text="원격 DB 저장", command=self.save_product_ui, bg="#4CAF50", fg="white", width=15).pack(side="left", padx=5)
        tk.Button(btn_frame, text="DB에서 삭제", command=self.delete_product_ui, bg="#f44336", fg="white", width=10).pack(side="left", padx=5)
        tk.Button(form_frame, text="🟢 상품 엑셀 파일 대량 등록 (DB 벌크 인서트)", command=self.import_product_excel, bg="#fff9c4", fg="#f57f17", font=("맑은 고딕", 10, "bold")).grid(row=3, column=0, columnspan=6, sticky="we", pady=5, padx=10)

        tree_frame = tk.Frame(tab3)
        tree_frame.pack(padx=15, pady=5, fill="both", expand=True)
        self.p_tree = ttk.Treeview(tree_frame, columns=("code", "w", "d", "h"), show="headings", height=13)
        self.p_tree.heading("code", text="옵션코드")
        self.p_tree.heading("w", text="가로(cm)")
        self.p_tree.heading("d", text="세로(cm)")
        self.p_tree.heading("h", text="높이(cm)")
        
        p_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.p_tree.yview)
        self.p_tree.configure(yscrollcommand=p_scrollbar.set)
        self.p_tree.pack(side="left", fill="both", expand=True)
        p_scrollbar.pack(side="right", fill="y")
        
        self.p_tree.bind("<ButtonRelease-1>", self.on_product_select)
        self.update_product_tree()

    def update_product_tree(self):
        self.p_tree.delete(*self.p_tree.get_children())
        self.products = load_products()
        search_keyword = self.p_search_var.get().strip().upper()
        for code, dims in self.products.items():
            if not search_keyword or search_keyword in code.upper():
                self.p_tree.insert("", tk.END, values=(code, dims[0], dims[1], dims[2]))

    def on_product_select(self, event):
        sel = self.p_tree.selection()
        if sel:
            vals = self.p_tree.item(sel[0], "values")
            self.ent_p_code.delete(0, tk.END); self.ent_p_code.insert(0, vals[0])
            self.ent_p_w.delete(0, tk.END); self.ent_p_w.insert(0, vals[1])
            self.ent_p_d.delete(0, tk.END); self.ent_p_d.insert(0, vals[2])
            self.ent_p_h.delete(0, tk.END); self.ent_p_h.insert(0, vals[3])

    def save_product_ui(self):
        code = self.ent_p_code.get().strip()
        if not code: return
        try:
            w, d, h = float(self.ent_p_w.get() or 0), float(self.ent_p_d.get() or 0), float(self.ent_p_h.get() or 0)
            save_or_update_product(code, (w, d, h))
            self.update_product_tree()
            self.ent_p_code.delete(0, tk.END); self.ent_p_w.delete(0, tk.END)
            self.ent_p_d.delete(0, tk.END); self.ent_p_h.delete(0, tk.END)
        except ValueError: 
            messagebox.showerror("오류", "규격은 숫자만 입력 가능합니다.")

    def delete_product_ui(self):
        sel = self.p_tree.selection()
        if not sel: return
        if messagebox.askyesno("삭제 확인", "선택한 상품을 DB에서 삭제하시겠습니까?"):
            code = self.p_tree.item(sel[0], "values")[0]
            delete_product_from_db(code)
            self.update_product_tree()
            self.ent_p_code.delete(0, tk.END); self.ent_p_w.delete(0, tk.END)
            self.ent_p_d.delete(0, tk.END); self.ent_p_h.delete(0, tk.END)

    def import_product_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
        if not file_path: return
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True): 
                if row[0] is not None:
                    save_or_update_product(str(row[0]).strip(), (float(row[1] or 0), float(row[2] or 0), float(row[3] or 0)))
            self.update_product_tree()
            messagebox.showinfo("완료", "원격 DB로 상품 엑셀 업로드 성공!")
        except Exception as e: 
            messagebox.showerror("오류", f"엑셀 처리 중 에러가 발생했습니다:\n{str(e)}")

    def create_tab4(self):
        tab4 = tk.Frame(self.notebook)
        self.notebook.add(tab4, text=" 4번 탭 (조합별 박스 제어) ")
        
        search_frame = tk.LabelFrame(tab4, text=" 🔍 상품코드 조합 검색 ")
        search_frame.pack(padx=15, pady=5, fill="x")
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *args: self.update_combo_tree())
        tk.Entry(search_frame, textvariable=self.search_var, width=35).pack(side="left", padx=10, pady=5)
        
        form_frame = tk.LabelFrame(tab4, text=" 조합별 지정 박스 등록 / 수정 (원격 DB 저장) ")
        form_frame.pack(padx=15, pady=5, fill="x")
        
        tk.Label(form_frame, text="상품코드 조합 (+ 구분):").grid(row=0, column=0, padx=5, pady=5)
        self.ent_c_key = tk.Entry(form_frame, width=40); self.ent_c_key.grid(row=0, column=1, padx=5, pady=5)
        tk.Label(form_frame, text="지정 박스명:").grid(row=1, column=0, padx=5, pady=5)
        self.ent_c_box = tk.Entry(form_frame, width=20); self.ent_c_box.grid(row=1, column=1, padx=5, pady=5)
        
        btn_frame = tk.Frame(form_frame)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=5)
        tk.Button(btn_frame, text="원격 DB 저장", command=self.save_combo_ui, bg="#4CAF50", fg="white", width=15).pack(side="left", padx=5)
        tk.Button(btn_frame, text="DB에서 삭제", command=self.delete_combo_ui, bg="#f44336", fg="white", width=10).pack(side="left", padx=5)
        
        tk.Button(
            form_frame, 
            text="🟢 조합 엑셀 파일 대량 등록 (A열: 조합 / B열: 박스)", 
            command=self.import_combo_excel, 
            bg="#e8f5e9", fg="#2e7d32", font=("맑은 고딕", 10, "bold")
        ).grid(row=3, column=0, columnspan=2, sticky="we", pady=5, padx=10)
        
        tree_frame = tk.Frame(tab4)
        tree_frame.pack(padx=15, pady=5, fill="both", expand=True)
        self.c_tree = ttk.Treeview(tree_frame, columns=("combo", "box"), show="headings", height=12)
        self.c_tree.heading("combo", text="상품 옵션코드 조합 목록")
        self.c_tree.heading("box", text="지정 박스 명칭")
        self.c_tree.column("combo", width=420)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.c_tree.yview)
        self.c_tree.configure(yscrollcommand=scrollbar.set)
        self.c_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.c_tree.bind("<ButtonRelease-1>", self.on_combo_select)
        self.update_combo_tree()

    def update_combo_tree(self):
        self.c_tree.delete(*self.c_tree.get_children())
        self.combos = load_combination_boxes()
        search_keyword = self.search_var.get().strip().upper()
        for combo_key, box_name in self.combos.items():
            if not search_keyword or search_keyword in combo_key.upper():
                self.c_tree.insert("", tk.END, values=(combo_key, box_name))

    def on_combo_select(self, event):
        sel = self.c_tree.selection()
        if sel:
            vals = self.c_tree.item(sel[0], "values")
            self.ent_c_key.delete(0, tk.END); self.ent_c_key.insert(0, vals[0])
            self.ent_c_box.delete(0, tk.END); self.ent_c_box.insert(0, vals[1])

    def save_combo_ui(self):
        raw_key = self.ent_c_key.get().strip()
        box_name = self.ent_c_box.get().strip()
        if raw_key and box_name:
            code_list = [c.strip() for c in raw_key.replace(',', '+').split('+') if c.strip()]
            combo_key = "+".join(sorted(code_list))
            save_or_update_combo(combo_key, box_name)
            self.update_combo_tree()
            self.ent_c_key.delete(0, tk.END); self.ent_c_box.delete(0, tk.END)

    def delete_combo_ui(self):
        sel = self.c_tree.selection()
        if not sel: return
        if messagebox.askyesno("삭제 확인", "선택한 조합을 DB에서 삭제하시겠습니까?"):
            combo_key = self.c_tree.item(sel[0], "values")[0]
            delete_combo_from_db(combo_key)
            self.update_combo_tree()
            self.ent_c_key.delete(0, tk.END); self.ent_c_box.delete(0, tk.END)

    def import_combo_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
        if not file_path: return
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            
            for row in sheet.iter_rows(min_row=2, values_only=True): 
                raw_combo = row[0]
                box_name = row[1]
                
                if raw_combo is not None and box_name is not None:
                    raw_combo_str = str(raw_combo).strip()
                    box_name_str = str(box_name).strip()
                    
                    if raw_combo_str and box_name_str:
                        code_list = [c.strip() for c in raw_combo_str.replace(',', '+').split('+') if c.strip()]
                        combo_key = "+".join(sorted(code_list))
                        save_or_update_combo(combo_key, box_name_str)
                        
            self.update_combo_tree()
            messagebox.showinfo("완료", "원격 DB로 지정 박스 조합 엑셀 업로드 성공!")
        except Exception as e: 
            messagebox.showerror("오류", f"엑셀 업로드 실패: {e}")

    # -------------------------------------------------------------------------
    # 5번 탭: 3D 미리보기 및 조합 등록 기능
    # -------------------------------------------------------------------------
    def create_tab5(self):
        tab5 = tk.Frame(self.notebook)
        self.notebook.add(tab5, text=" 5번 탭 (3D 미리보기 및 상자 등록) ")
        
        if not MATPLOTLIB_AVAILABLE:
            tk.Label(tab5, text="\n⚠️ 3D 시각화를 위해 matplotlib 라이브러리가 필요합니다.\n\n터미널에서 'pip install matplotlib'를 실행 후 프로그램을 재시작해주세요.", font=("맑은 고딕", 12), fg="red").pack(pady=50)
            return
            
        self.tab5_item_codes = []
        self.current_t5_combo_key = None
        self.current_t5_box_name = None

        # 왼쪽 프레임 (컨트롤 패널)
        left_frame = tk.Frame(tab5, width=280, bg="#f5f5f5")
        left_frame.pack(side="left", fill="y", padx=10, pady=10)
        left_frame.pack_propagate(False) 
        
        # 오른쪽 프레임 (3D 뷰어)
        right_frame = tk.Frame(tab5, bg="white", relief="solid", bd=1)
        right_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)

        # 컨트롤 패널 요소 배치
        tk.Label(left_frame, text="옵션코드 입력:", bg="#f5f5f5", font=("맑은 고딕", 10, "bold")).pack(pady=(10, 0), anchor="w", padx=10)
        input_frame = tk.Frame(left_frame, bg="#f5f5f5")
        input_frame.pack(fill="x", padx=10, pady=5)
        
        self.ent_t5_code = tk.Entry(input_frame, width=20)
        self.ent_t5_code.pack(side="left", fill="x", expand=True)
        self.ent_t5_code.bind('<Return>', lambda e: self.add_t5_item())
        tk.Button(input_frame, text="추가", command=self.add_t5_item, bg="#03a9f4", fg="white").pack(side="right", padx=(5, 0))

        tk.Label(left_frame, text="현재 추가된 상품 목록:", bg="#f5f5f5").pack(pady=(15, 0), anchor="w", padx=10)
        self.lb_t5_items = tk.Listbox(left_frame, height=8)
        self.lb_t5_items.pack(fill="x", padx=10, pady=5)
        
        tk.Button(left_frame, text="목록 초기화", command=self.clear_t5_items, bg="#9e9e9e", fg="white").pack(fill="x", padx=10, pady=5)
        
        tk.Label(left_frame, text="-"*40, bg="#f5f5f5").pack(pady=10)
        tk.Button(left_frame, text="▶ 3D 시뮬레이션 및 상자 추천", command=self.run_3d_simulation, bg="#3f51b5", fg="white", font=("맑은 고딕", 10, "bold"), height=2).pack(fill="x", padx=10)

        result_frame = tk.LabelFrame(left_frame, text=" 결과 정보 ", bg="#f5f5f5")
        result_frame.pack(fill="x", padx=10, pady=15)
        self.lbl_t5_result = tk.Label(result_frame, text="-", font=("맑은 고딕", 11, "bold"), fg="#e91e63", bg="#f5f5f5")
        self.lbl_t5_result.pack(pady=10)
        
        tk.Button(left_frame, text="✅ 이 조합을 4번 탭(DB)에 등록", command=self.register_combo_from_tab5, bg="#4CAF50", fg="white", font=("맑은 고딕", 10, "bold")).pack(fill="x", padx=10, pady=5)

        # 3D 뷰어 셋업
        self.fig = plt.Figure(figsize=(5, 5), dpi=100)
        self.ax = self.fig.add_subplot(111, projection='3d')
        self.canvas = FigureCanvasTkAgg(self.fig, master=right_frame)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)

    def add_t5_item(self):
        code = self.ent_t5_code.get().strip()
        if not code: return
        
        products = load_products()
        if code not in products:
            messagebox.showwarning("상품 없음", f"'{code}' 상품은 3번 탭(DB)에 등록되어 있지 않습니다.\n크기 정보를 먼저 등록해주세요.")
            return
            
        self.lb_t5_items.insert(tk.END, code)
        self.tab5_item_codes.append(code)
        self.ent_t5_code.delete(0, tk.END)

    def clear_t5_items(self):
        self.lb_t5_items.delete(0, tk.END)
        self.tab5_item_codes.clear()
        self.lbl_t5_result.config(text="-")
        self.current_t5_combo_key = None
        self.current_t5_box_name = None
        self.ax.clear()
        self.canvas.draw()

    def draw_box_wireframe(self, w, d, h):
        points = [
            [(0,0,0), (w,0,0)], [(0,0,0), (0,d,0)], [(0,0,0), (0,0,h)],
            [(w,0,0), (w,d,0)], [(w,0,0), (w,0,h)],
            [(0,d,0), (w,d,0)], [(0,d,0), (0,d,h)],
            [(0,0,h), (w,0,h)], [(0,0,h), (0,d,h)],
            [(w,d,0), (w,d,h)],
            [(w,0,h), (w,d,h)],
            [(0,d,h), (w,d,h)]
        ]
        for p1, p2 in points:
            self.ax.plot([p1[0], p2[0]], [p1[1], p2[1]], [p1[2], p2[2]], color='red', linestyle='--', alpha=0.5, linewidth=1.5)

    def run_3d_simulation(self):
        if not self.tab5_item_codes:
            messagebox.showinfo("안내", "상품 옵션코드를 먼저 추가해주세요.")
            return
            
        products = load_products()
        items_with_ids = []
        for code in self.tab5_item_codes:
            if code in products:
                dims = products[code]
                items_with_ids.append((code, dims[0], dims[1], dims[2]))
            else:
                messagebox.showerror("오류", f"'{code}' 상품 크기 정보가 없습니다.")
                return
                
        # 새로운 Best-Fit 3D 팩킹 알고리즘 연산
        best_box_name, best_box_dims, layout = find_best_box_and_layout(items_with_ids, BOX_CANDIDATES)
        
        self.current_t5_combo_key = "+".join(sorted(self.tab5_item_codes))
        self.current_t5_box_name = best_box_name if best_box_dims else None
        
        if best_box_dims:
            self.lbl_t5_result.config(text=f"추천 상자: {best_box_name}\n({best_box_dims[0]}x{best_box_dims[1]}x{best_box_dims[2]} cm)")
        else:
            self.lbl_t5_result.config(text="⚠️ 규격초과\n(맞는 상자 없음)")
            
        # 3D 화면 지우기 및 그리기 시작
        self.ax.clear()
        
        if best_box_dims:
            box_w, box_d, box_h = best_box_dims
            self.draw_box_wireframe(box_w, box_d, box_h)
            
            colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b', '#e377c2']
            for i, space in enumerate(layout):
                px, py, pz, iw, id_, ih, item_code = space
                c = colors[i % len(colors)]
                self.ax.bar3d(px, py, pz, iw, id_, ih, alpha=0.6, color=c, edgecolor='black', linewidth=0.5)
                # 상품 코드 라벨 추가
                self.ax.text(px + iw/2, py + id_/2, pz + ih/2, item_code, color='black', ha='center', va='center', fontsize=8)

            max_dim = max(best_box_dims)
            self.ax.set_xlim([0, max_dim])
            self.ax.set_ylim([0, max_dim])
            self.ax.set_zlim([0, max_dim])
            
        self.ax.set_xlabel('Width (cm)')
        self.ax.set_ylabel('Depth (cm)')
        self.ax.set_zlabel('Height (cm)')
        self.ax.set_title("3D Box Packing Preview")
        
        self.canvas.draw()

    def register_combo_from_tab5(self):
        if not hasattr(self, 'current_t5_combo_key') or not self.current_t5_combo_key:
            messagebox.showwarning("경고", "먼저 3D 시뮬레이션을 실행하세요.")
            return
        if not self.current_t5_box_name:
            messagebox.showwarning("경고", "추천된 상자가 없어 등록할 수 없습니다.")
            return
            
        save_or_update_combo(self.current_t5_combo_key, self.current_t5_box_name)
        self.update_combo_tree()
        messagebox.showinfo("등록 완료", f"조합: [{self.current_t5_combo_key}]\n상자: {self.current_t5_box_name}\n\n4번 탭(조합별 지정 박스) DB에 성공적으로 등록되었습니다.")

if __name__ == "__main__":
    root = TkinterDnD.Tk()
    app = IntegratedBoxApp(root)
    root.mainloop()